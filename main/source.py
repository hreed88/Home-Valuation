from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import functools
import math
import multiprocessing
import pickle
import smtplib
import ssl
import sys
import threading
from time import sleep
import time
from typing import KeysView
from urllib.request import urlopen
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import NoSuchElementException
import tkinter as tk
from multiprocessing import shared_memory

class homeVal(multiprocessing.Process):

    def __init__(self) -> None:
        #multiprocessing.Process.__init__(self)
        self.data = []#multiprocessing.Array('c',0)
        self.driver = None
        self.dataRetrived = False #multiprocessing.Value('b', False)
        pass

    def getData(self, bookName):
        data = openpyxl.load_workbook(bookName)
        data1 = data.active

        myData = []

        for i in range(1, data1.max_row):
            temp = [-1]*7
            j = 0
            thisBreak = False
            for col in data1.iter_cols(0, 7):
                if col[i].value == None or col[i].value == "unknown" or col[i].value == "Unknown":
                    thisBreak = True
                temp[j] = col[i].value
                j+=1
            if(thisBreak):
                continue
            myData.insert(0,temp)
        
        result = []
        [result.append(x) for x in myData if x not in result]
        #Returns arrays in the format[First Name, Last Name, State, Address, Zip, Email]
        #return result
        self.data = result#multiprocessing.Array('c',result)
        self.dataRetrived= True#.value=True
        pass

    def parseData(self):
        #Open browser
        self.driver = webdriver.Chrome()
        #self.driver.get(#Put url here and uncomment line)
        
        
        self.driver.find_element(By.ID, "warning-login-btn").click()
        user = self.driver.find_element(By.ID, "userNameInput")
        password = self.driver.find_element(By.ID,"passwordInput")
        #Login
        #user.send_keys("#Put user name here and uncomment")
        #password.send_keys("#Put Password here and uncomment")

        self.driver.find_element(By.ID,"submitButton").click()
        sleep(10)
        
        newData = []
        print("operating")
        avg = 1
        for i in range(5):
            
            start = time.process_time()
            #self.driver.get("#Put same url here as used in line 67")
            sleep(1)
            address = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.ID, "street-address")))
            city = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.ID, "city")))
            zip = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.ID, "zip-code")))
            state = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.ID, "state")))
            selectState = Select(state)
            address.send_keys(self.data[i][3])
            city.send_keys(self.data[i][5])
            zip.send_keys(self.data[i][4])
            selectState.select_by_value(self.data[i][2])

            self.driver.find_element(By.ID,"house-canary-verify-address").click()
            sleep(1)
            temp = self.check_exists_by_xpath(self.driver,"//button[@class = 'btn alt get-value']")
            if(temp):
                self.driver.find_element(By.XPATH, "//button[@class = 'btn alt get-value']").click()
    
            try:
                value = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.XPATH, "//p[@class = 'estimated-cost bold']"))).text
                link = self.driver.find_element(By.XPATH, "//a[@class = 'btn secondary large view-interactive-report-button']").get_attribute('href')
            except:
                print("Element ", self.data[i] ," Removed, Couldnt retrive data")
                continue
            print(link)
            print(value)
            
            #Update array[index] if values were found 
            self.data[i].append(link)
            self.data[i].append(value)
            #Add to new array
            newData.append(self.data[i])
            #Calculated avg time remaining
            end = time.process_time()
            avg += end-start
            avg /= 2
            print(self.data[i])
            print("Current itr", i)
            print(len(self.data))
            print("Estimated minutes remaining: ", math.floor(avg*(len(self.data) - i)))
            
            
        self.data = newData

        
    def check_exists_by_xpath(self, thisDriver ,xpath):
        print(xpath)
        try:
            thisDriver.find_element(By.XPATH, xpath)
        except NoSuchElementException:
            return False
        return True

    def email(self):
        port = 465
        password = input("Enter Password:")
        #recip = ["#Put recipients here"]#Note used for testing
        context = ssl.create_default_context()
        
        with smtplib.SMTP_SSL("smtp.gmail.com", port, context=context) as server:
            #server.login("Put email to send out to recipients", password)
            
            for i in range(5):
                if(len(self.data[i]) != 9):
                    print("cant send")
                    continue
                msg = MIMEMultipart()
                msg['Subject'] = "Your Home's Valuation: Unlocking Insights for Your Future! "
                msg['From'] = #put email to send out to recipients here
                msg['to'] = self.data[i][6]
                message = """<p>Will be sent to""" + self.data[i][6] + """<br>Dear """ + str(self.data[i][0]) +""",<br><br> 

    I hope this message finds you well and in the holiday spirits!<br><br>  

    As your trusted Mortgage Advisor and now that I am a Real Estate Agent as well, I'm thrilled to share<br>some valuable insights regarding your home's current value in this vibrant market.<br><br> 

    The real estate landscape in your neighborhood has been dynamic, and keeping you informed about the<br>precise worth of your property is a priority. Whether it's for strategic planning, investment prospects, or<br>simply staying updated on your asset's value, this information is pivotal.<br><br> 

    I've meticulously compiled a comprehensive report that mirrors the assessment an underwriter uses to<br>verify a home's value, adding an extra layer of assurance beyond the typical appraisal process. It's a neat<br>tool and I want to share this information! Today, based on the latest market data, your home located at<br>"""+ str(self.data[i][3]) + """ boasts an estimated value of: 
    
    <br><br><b style="font-size:20px;">"""+ str(self.data[i][8]) + """</b><br><br>

    Curious to explore the detailed report for yourself? Just click <a href=" """+ self.data[i][7] +""" ">HERE</a> and once opened, you'll have the option <br>to generate a PDF version of the report. What's particularly valuable within this report is the insight <br>into the high and low-value range along with the average neighborhood rents. It's a fantastic gauge <br>if you've ever pondered becoming a landlord - perhaps utilizing your home's equity for a down payment <br>on a new property and transforming your current home into a rental gem!<br><br> 

    Feel free to delve into this report at your convenience, and should you have any queries or wish to<br>discuss this valuation further, I'm here to assist.<br><br> 

    Warm regards,</p><br><br>"""
                #message = MIMEText(message, 'plain', 'utf-8')
                message = MIMEText(message, 'html')
                temp = MIMEText('<img src="cid:image1" width=500 height=auto>', 'html')
                attachment = open("footer.png",'rb')
                img_data = attachment.read()
                image = MIMEImage(img_data)
                image.add_header('Content-ID','<image1>')
                footer = MIMEText("""<p>NMLS 278757 MLO 100015425</p><p>8300 E Maplewood Ave #100 Greenwood Village, CO 80111  EC100054186</p><small style="font-size:10px;">PHISHING SCAMS CONNTINUE – PLEASE BEWARE, the bad guys never cease to let a Crisis go to waste.  Please be cautious of opening emails disguised as Coronavirus or Covid-19.<br> The virus is not related to your real estate transaction.  If you get a suspicious email - please call us immediately. Remember, the best course of action is to not open links you are not expecting.</small><br><br><i style="font-size:10px;">CONFIDENTIALITY NOTICE: This email message and any attachments are for the exclusive use of the intended recipient(s) and may contain<br>confidential and privileged information. Any unauthorized review, use, disclosure or distribution is prohibited. If you are not the intended recipient,<br> please contact the sender by reply email and destroy all copies of the original message along with any attachments, from your computer system.<br>If you are the intended recipient, please be advised that the content of this message is subject to access, review and disclosure by the sender's <br>Email System Administrator.</i>""",'html')
                msg.attach(message)
                msg.attach(temp)
                msg.attach(image)
                msg.attach(footer)
                
                #msg.attach(image)
                server.sendmail("Your email", self.data[i][6], msg.as_string())
                print("Message sent to " + self.data[i][6])
        pass

    def isOpen(driver, temp2):
        try:
            driver.switch_to.window(temp2)
            return True
        except:
            return False
"""
def startMultHome():
    shm = shared_memory.SharedMemory(create=True, size=(sys.getsizeof(myVal)))
    temp = homeVal(buffer = shm.buf)
    temp.data = myVal.data
    temp.driver = myVal.driver
    temp.dataRetrived = myVal.dataRetrived
    p = multiprocessing.Process(target=myVal.parseData)
    p.start()
    shm.close()

def startMultData():
    
    temp = multiprocessing.Value('b',myVal.dataRetrived.value)
    p = multiprocessing.Process(target=myVal.getData,args=("book1.xlsx",))
    p.start()
    p.join()

def startMultEmail():

    p = multiprocessing.Process(target=myVal.email)
    p.start()
"""
def test1():
    threading.Thread(target=myVal.getData('book1.xlsx')).start()

def test2():
    threading.Thread(target=myVal.parseData()).start()

def test3():
    threading.Thread(target=myVal.email()).start()

def runFunc(thisRun,currData):
        while(thisRun):
            window.update_idletasks()
            window.update()
            if(myVal.dataRetrived):
                currData = myVal.data
            try:
                window.winfo_exists()
            except:    
                thisRun = False
        pass

if __name__ == '__main__':
        myVal = homeVal()
        #multiprocessing.Value('b', False))
        #Get old data if thats what were working with
        with open('oldData.txt','rb') as fp:
            try:
                currData = pickle.load(fp)
                myVal.data = currData
            except:
                print("No data to load from")
        
        window = tk.Tk()
        window.title("Home Valuation Tool")
        window.geometry("800x400")
        
        #greeting = tk.Label(text="Hello, tkinter", foreground="white", background="black")
        #greeting.pack()

        isRunning = True
        getDataButt = tk.Button(text="New Data", width=25, height=5,command=test1)#lambda:myVal.getData('book1.xlsx'))
        getHomeVal = tk.Button(text="Get Home Values", width=25, height=5, command=test2)
        sendEmail = tk.Button(text="Send Emails", width=25, height=5,command=test3)
        #exit = tk.Button(text="Exit", width=15,height=5,command= (isRunning == False))
        print(window.winfo_height())
        #exit.place(x=350,y=200)
        getDataButt.place(x=100,y=125)
        getHomeVal.place(x=300,y=125)
        sendEmail.place(x=500,y=125)
        threading.Thread(target=runFunc(isRunning,currData))
        with open('oldData.txt', 'wb')as fp:
            pickle.dump(currData, fp)
